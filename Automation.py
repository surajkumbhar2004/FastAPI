from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, HTMLResponse
import pandas as pd
import os
from io import BytesIO
from rapidfuzz import process, fuzz  
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Define the output folder path for saving the processed result
os.makedirs("output_folder", exist_ok=True)

app = FastAPI()

@app.get("/", response_class=HTMLResponse)
async def root():
    return """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Upload Files</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            background-color: #f0f8ff; /* Light blue background */
            margin: 0;
            padding: 20px;
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100vh;
            box-sizing: border-box;
        }

        .container {
            text-align: center;
        }

        h1 {
            color: #444;
            font-size: 28px;
            margin-bottom: 20px;
        }

        form {
            background: #ffffff;
            padding: 25px;
            border-radius: 10px;
            box-shadow: 0 4px 15px rgba(0, 0, 0, 0.1);
            width: 400px;
            text-align: left;
        }

        label {
            display: block;
            font-size: 16px;
            margin-bottom: 8px;
            color: #666;
        }

        input[type="file"] {
            margin-bottom: 20px;
            padding: 12px;
            border: 1px solid #ccc;
            border-radius: 6px;
            width: 100%;
            box-sizing: border-box;
        }

        button {
            background-color: #4682b4; /* Darker sky blue */
            color: white;
            border: none;
            padding: 12px 18px;
            border-radius: 8px;
            cursor: pointer;
            transition: background-color 0.3s;
            width: 100%;
            font-size: 16px;
        }

        button:hover {
            background-color: #3a6b92; /* Slightly darker on hover */
        }

        #loading {
            display: none;
            margin-top: 20px;
        }

        .hourglass {
            width: 24px;
            height: 24px;
            border: 6px solid transparent;
            border-top: 6px solid #4682b4;
            border-bottom: 6px solid #87ceeb;
            border-radius: 50%;
            animation: spin 1s linear infinite;
            margin: 0 auto;
        }

        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }

        @media (max-width: 450px) {
            form {
                width: 90%;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Upload Files</h1>
        <form id="uploadForm" enctype="multipart/form-data">
            <label for="file1">Input Source File:</label>
            <input type="file" name="file1" id="file1" accept=".xls,.xlsx" required>
            <label for="file2">Input Target File:</label>
            <input type="file" name="file2" id="file2" accept=".xls,.xlsx" required>
            <button type="submit">Upload Files</button>
            <div id="loading">
                <div class="hourglass"></div>
                <p>Processing...</p>
            </div>
        </form>
    </div>

    <script>
        document.getElementById("uploadForm").addEventListener("submit", async function(event) {
            event.preventDefault();  // Prevent normal form submission
            const formData = new FormData(event.target);

            document.getElementById("loading").style.display = "block";  // Show the loading animation

            try {
                const response = await fetch("/uploadfile/", {
                    method: "POST",
                    body: formData
                });

                if (response.ok) {
                    const blob = await response.blob();
                    const url = window.URL.createObjectURL(blob);

                    // Create a temporary link to download the file
                    const a = document.createElement("a");
                    a.href = url;
                    a.download = "output_file.xlsx";  // Updated filename
                    document.body.appendChild(a);
                    a.click();

                    window.URL.revokeObjectURL(url);  // Clean up the URL object
                } else {
                    alert("Error processing the file.");
                }
            } catch (error) {
                console.error("Error:", error);
                alert("An unexpected error occurred.");
            } finally {
                document.getElementById("loading").style.display = "none";  // Hide the loading animation
            }
        });
    </script>
</body>
</html>

    """


@app.post("/uploadfile/")
async def process_files(file1: UploadFile = File(...), file2: UploadFile = File(...)):
    try:
        # Read and process the "Extract Maconomy" file
        contents1 = await file1.read()
        df1 = pd.read_excel(BytesIO(contents1))

      
        # cleaning Extract Maconomy File
        # 1. Filter rows based on Vendor No.
        df1 = df1[df1['Vendor No.'].astype(str).str.startswith(('IC', 'STG'))]

        # 2. Clean the 'FEIN Number' column by removing spaces and dashes
        df1['FEIN Number'] = df1['FEIN Number'].astype(str).str.replace(' ', '').str.replace('-', '', regex=False)

        # 3. Clean the 'Vendor Name' column
        df1['Vendor Name'] = df1['Vendor Name'].astype(str)
        df1['Vendor Name'] = df1['Vendor Name'].str.replace(',', '', regex=False)  # Remove commas
        df1['Vendor Name'] = df1['Vendor Name'].str.replace('.', '', regex=False)  # Remove periods
        df1['Vendor Name'] = df1['Vendor Name'].str.replace('/', '', regex=False)  # Remove slashes
        df1['Vendor Name'] = df1['Vendor Name'].str.strip()  # Trim whitespace

        # Read and process the "MD3 Vendors" file
        contents2 = await file2.read()
        df2 = pd.read_excel(BytesIO(contents2), sheet_name='Vendor Template')

        # Set header and clean df2
        df2.columns = df2.iloc[0]
        df2 = df2.drop(index=[0, 1, 2, 3]).reset_index(drop=True)

        df2['Vendor Name'] = df2['Vendor Name'].astype(str)
        df2['Vendor Name'] = df2['Vendor Name'].str.replace(',', '').str.replace('.', '').str.replace('/', '')
        df2['Vendor Name'] = df2['Vendor Name'].str.replace(r'(?i)\b(LLC|CORP|INC|LTD|dba|org)\b|[\(\)\-]', '', regex=True).str.strip()

        # Define the function for matching
        def find_vendor(parentvendor, fein, vendor_name, df1, threshold=75):
            if fein and not pd.isna(fein):
                fein_matches = df1[df1['FEIN Number'] == fein]
                if len(fein_matches) == 1:
                    return fein_matches['Vendor No.'].values[0], "FEIN Match", None
                elif len(fein_matches) > 1:
                    vendor_no = fein_matches['Vendor No.'].values[0]
                    duplicate_indices = fein_matches.index.tolist()[1:]
                    return vendor_no, "FEIN Match with Duplicates", duplicate_indices

            vendor_names = df1['Vendor Name'].tolist()
            close_matches = process.extract(vendor_name, vendor_names, scorer=fuzz.ratio, limit=5)
            close_matches = [match for match in close_matches if match[1] >= threshold]

            if len(close_matches) == 1:
                closest_vendor = close_matches[0][0]
                vendor_no = df1.loc[df1['Vendor Name'] == closest_vendor, 'Vendor No.'].values[0]
                return vendor_no, "Close Match", None

            elif len(close_matches) > 1:
                close_matches.sort(key=lambda x: x[1], reverse=True)
                parent_vendor_matches = [
                    match for match in close_matches
                    if df1[df1['Vendor Name'] == match[0]]['Parent Vendor'].values[0] == parentvendor
                ]
                if parent_vendor_matches:
                    best_match = parent_vendor_matches[0][0]
                    vendor_no = df1.loc[df1['Vendor Name'] == best_match, 'Vendor No.'].values[0]
                    return vendor_no, "ParentVendor Match", None

                best_match = close_matches[0][0]
                vendor_no = df1.loc[df1['Vendor Name'] == best_match, 'Vendor No.'].values[0]
                duplicate_names = [match[0] for match in close_matches[1:]]
                duplicate_indices = df1[df1['Vendor Name'].isin(duplicate_names)].index.tolist()
                return vendor_no, "Highest Score Match", duplicate_indices

            return None, "No Match", None

        # Apply the matching function
        df2['Vendor No.'], df2['Match Status'], df2['Duplicate Rows'] = zip(
            *df2.apply(lambda row: find_vendor(row['Parent Vendor'], row['FEIN #'], row['Vendor Name'], df1), axis=1)
        )

        # Save the processed result
        output_file = os.path.join("output_folder", 'output.xlsx')
        df2.to_excel(output_file, index=False)

        # Load workbook for formatting using openpyxl
        wb = load_workbook(output_file)
        ws = wb.active

        # Check if there are no sheets or active sheet
        if ws is None or ws.max_row <= 1:
            raise HTTPException(status_code=500, detail="No data rows available in the sheet.")

        # Define color fills
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Use pandas to identify column indices for "Vendor Name" and "Match Status"
        vendor_name_col = df2.columns.get_loc("Vendor Name")
        if isinstance(vendor_name_col, int):
           vendor_name_col += 1  
        else:
            raise TypeError("Expected an integer index, but got: " + str(type(vendor_name_col)))

        match_status_col = df2.columns.get_loc("Match Status") 
        if isinstance(match_status_col, int):
           match_status_col += 1  
        else:
            raise TypeError("Expected an integer index, but got: " + str(type(match_status_col)))


        # Apply color fills based on match status
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            vendor_name_cell = row[vendor_name_col - 1]  # Adjusting for zero-based index
            match_status_cell = row[match_status_col - 1]

            if match_status_cell.value == "Close Match":
                vendor_name_cell.fill = green_fill
            elif match_status_cell.value == "ParentVendor Match":
                vendor_name_cell.fill = orange_fill
            elif match_status_cell.value == "No Match":
                vendor_name_cell.fill = red_fill

        # Save the workbook with applied styles
        wb.save(output_file)

        return FileResponse(output_file, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename='output.xlsx')

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="127.0.0.1", port=8080)
